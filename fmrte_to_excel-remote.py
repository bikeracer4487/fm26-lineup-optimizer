"""
FMRTE to Excel Automation Script
---------------------------------
Automates the process of copying player data from Football Manager Real Time Editor (FMRTE)
into an Excel spreadsheet for the FM26 Lineup Optimizer.

This version uses pyautogui for direct mouse/keyboard control since FMRTE has custom UI
controls that don't work well with pywinauto's accessibility APIs.

Author: Doug Mason (2025)
License: MIT
"""

import time
import subprocess
import sys
import argparse
from pathlib import Path
import ctypes
import ctypes.wintypes

# Set DPI awareness BEFORE importing pyautogui
# This is critical for high-DPI displays
try:
    # Try Windows 10+ method first
    ctypes.windll.shcore.SetProcessDpiAwareness(2)  # PROCESS_PER_MONITOR_DPI_AWARE
except Exception:
    try:
        # Fallback for older Windows
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

# Try to import required libraries
try:
    import pyautogui
    pyautogui.FAILSAFE = True  # Move mouse to corner to abort
    pyautogui.PAUSE = 0.05  # Small pause between pyautogui commands
except ImportError:
    print("ERROR: pyautogui not installed. Run: pip install pyautogui")
    sys.exit(1)

try:
    import pyperclip
except ImportError:
    print("ERROR: pyperclip not installed. Run: pip install pyperclip")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import win32gui
    import win32con
    import win32api
except ImportError:
    print("ERROR: pywin32 not installed. Run: pip install pywin32")
    sys.exit(1)


# Configuration
EXCEL_FILE_PATH = Path(r"C:\GitHub\fm26-lineup-optimizer\FM26 Players.xlsx")
TARGET_SHEET = "Paste Full"
FMRTE_WINDOW_TITLE = "FMRTE"
TABS = ["Brixham", "Brixham U21", "Brixham U18"]
ACTION_DELAY = 0.4  # Seconds to wait between actions
COPY_DELAY = 0.8    # Extra time to wait for clipboard operations


def move_mouse_raw(x, y):
    """
    Move mouse using SendInput API (more reliable than mouse_event).
    
    Args:
        x, y: Screen coordinates
        
    Returns:
        tuple: (new_x, new_y) actual position after move
    """
    # Define structures for SendInput
    class MOUSEINPUT(ctypes.Structure):
        _fields_ = [
            ("dx", ctypes.c_long),
            ("dy", ctypes.c_long),
            ("mouseData", ctypes.c_ulong),
            ("dwFlags", ctypes.c_ulong),
            ("time", ctypes.c_ulong),
            ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
        ]
    
    class KEYBDINPUT(ctypes.Structure):
        _fields_ = [
            ("wVk", ctypes.c_ushort),
            ("wScan", ctypes.c_ushort),
            ("dwFlags", ctypes.c_ulong),
            ("time", ctypes.c_ulong),
            ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
        ]
    
    class HARDWAREINPUT(ctypes.Structure):
        _fields_ = [
            ("uMsg", ctypes.c_ulong),
            ("wParamL", ctypes.c_short),
            ("wParamH", ctypes.c_ushort)
        ]
    
    class INPUT_UNION(ctypes.Union):
        _fields_ = [
            ("mi", MOUSEINPUT),
            ("ki", KEYBDINPUT),
            ("hi", HARDWAREINPUT)
        ]
    
    class INPUT(ctypes.Structure):
        _fields_ = [
            ("type", ctypes.c_ulong),
            ("union", INPUT_UNION)
        ]
    
    # Get virtual screen dimensions
    SM_XVIRTUALSCREEN = 76
    SM_YVIRTUALSCREEN = 77  
    SM_CXVIRTUALSCREEN = 78
    SM_CYVIRTUALSCREEN = 79
    
    vscreen_x = ctypes.windll.user32.GetSystemMetrics(SM_XVIRTUALSCREEN)
    vscreen_y = ctypes.windll.user32.GetSystemMetrics(SM_YVIRTUALSCREEN)
    vscreen_width = ctypes.windll.user32.GetSystemMetrics(SM_CXVIRTUALSCREEN)
    vscreen_height = ctypes.windll.user32.GetSystemMetrics(SM_CYVIRTUALSCREEN)
    
    # Convert to normalized coordinates (0-65535)
    norm_x = int((x - vscreen_x) * 65535 / vscreen_width)
    norm_y = int((y - vscreen_y) * 65535 / vscreen_height)
    
    # Clamp
    norm_x = max(0, min(65535, norm_x))
    norm_y = max(0, min(65535, norm_y))
    
    # MOUSEEVENTF_MOVE = 0x0001
    # MOUSEEVENTF_ABSOLUTE = 0x8000
    # MOUSEEVENTF_VIRTUALDESK = 0x4000
    flags = 0x0001 | 0x8000 | 0x4000
    
    inp = INPUT()
    inp.type = 0  # INPUT_MOUSE
    inp.union.mi.dx = norm_x
    inp.union.mi.dy = norm_y
    inp.union.mi.mouseData = 0
    inp.union.mi.dwFlags = flags
    inp.union.mi.time = 0
    inp.union.mi.dwExtraInfo = None
    
    result = ctypes.windll.user32.SendInput(1, ctypes.byref(inp), ctypes.sizeof(inp))
    if result != 1:
        error = ctypes.get_last_error()
        print_debug(f"SendInput move failed: returned {result}, error {error}")
    
    time.sleep(0.05)
    
    # Get new position
    point = ctypes.wintypes.POINT()
    ctypes.windll.user32.GetCursorPos(ctypes.byref(point))
    return (point.x, point.y)


def click_raw():
    """
    Perform a mouse click using SendInput API.
    """
    class MOUSEINPUT(ctypes.Structure):
        _fields_ = [
            ("dx", ctypes.c_long),
            ("dy", ctypes.c_long),
            ("mouseData", ctypes.c_ulong),
            ("dwFlags", ctypes.c_ulong),
            ("time", ctypes.c_ulong),
            ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
        ]
    
    class KEYBDINPUT(ctypes.Structure):
        _fields_ = [
            ("wVk", ctypes.c_ushort),
            ("wScan", ctypes.c_ushort),
            ("dwFlags", ctypes.c_ulong),
            ("time", ctypes.c_ulong),
            ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
        ]
    
    class HARDWAREINPUT(ctypes.Structure):
        _fields_ = [
            ("uMsg", ctypes.c_ulong),
            ("wParamL", ctypes.c_short),
            ("wParamH", ctypes.c_ushort)
        ]
    
    class INPUT_UNION(ctypes.Union):
        _fields_ = [
            ("mi", MOUSEINPUT),
            ("ki", KEYBDINPUT),
            ("hi", HARDWAREINPUT)
        ]
    
    class INPUT(ctypes.Structure):
        _fields_ = [
            ("type", ctypes.c_ulong),
            ("union", INPUT_UNION)
        ]
    
    # Mouse down
    inp_down = INPUT()
    inp_down.type = 0  # INPUT_MOUSE
    inp_down.union.mi.dx = 0
    inp_down.union.mi.dy = 0
    inp_down.union.mi.mouseData = 0
    inp_down.union.mi.dwFlags = 0x0002  # MOUSEEVENTF_LEFTDOWN
    inp_down.union.mi.time = 0
    inp_down.union.mi.dwExtraInfo = None
    
    result = ctypes.windll.user32.SendInput(1, ctypes.byref(inp_down), ctypes.sizeof(inp_down))
    if result != 1:
        error = ctypes.get_last_error()
        print_debug(f"SendInput click down failed: returned {result}, error {error}")
    
    time.sleep(0.05)
    
    # Mouse up
    inp_up = INPUT()
    inp_up.type = 0  # INPUT_MOUSE
    inp_up.union.mi.dx = 0
    inp_up.union.mi.dy = 0
    inp_up.union.mi.mouseData = 0
    inp_up.union.mi.dwFlags = 0x0004  # MOUSEEVENTF_LEFTUP
    inp_up.union.mi.time = 0
    inp_up.union.mi.dwExtraInfo = None
    
    result = ctypes.windll.user32.SendInput(1, ctypes.byref(inp_up), ctypes.sizeof(inp_up))
    if result != 1:
        error = ctypes.get_last_error()
        print_debug(f"SendInput click up failed: returned {result}, error {error}")


def click_at(x, y):
    """
    Move mouse to coordinates and click using low-level Windows API.
    
    Args:
        x, y: Screen coordinates
        
    Returns:
        bool: True if successful
    """
    new_x, new_y = move_mouse_raw(x, y)
    
    # Check if move was successful (within 20 pixels tolerance)
    if abs(new_x - x) > 20 or abs(new_y - y) > 20:
        print_debug(f"Mouse move may have failed: target ({x}, {y}), actual ({new_x}, {new_y})")
    
    click_raw()
    return True


def print_divider():
    """Print a visual divider for terminal output."""
    print("=" * 70)


def print_status(message):
    """Print a status message."""
    print(f"[STATUS] {message}")


def print_error(message):
    """Print an error message."""
    print(f"[ERROR] {message}")


def print_debug(message):
    """Print a debug message."""
    print(f"[DEBUG] {message}")


def find_fmrte_window():
    """
    Find the FMRTE window handle and return its position/size.
    
    Returns:
        tuple: (hwnd, left, top, right, bottom) or None if not found
    """
    print_status("Searching for FMRTE window...")
    
    candidates = []  # Collect all matching windows
    
    def enum_callback(hwnd, extra):
        if win32gui.IsWindowVisible(hwnd):
            title = win32gui.GetWindowText(hwnd)
            class_name = win32gui.GetClassName(hwnd)
            
            # Must have FMRTE in the title
            if title and FMRTE_WINDOW_TITLE.upper() in title.upper():
                # Exclude terminals, consoles, browsers, editors
                excluded_classes = [
                    'WindowsTerminal', 'ConsoleWindowClass', 'CASCADIA_HOSTING_WINDOW_CLASS',
                    'Chrome_WidgetWin', 'MozillaWindowClass', 'Notepad', 'CabinetWClass',
                    'VSCODE', 'Cursor', 'mintty', 'PuTTY'
                ]
                excluded_titles = [
                    '.py', '.bat', 'cmd.exe', 'powershell', 'terminal', 'prompt',
                    'cursor', 'vscode', 'code', 'notepad', 'explorer'
                ]
                
                # Check class name exclusions
                is_excluded = any(exc.lower() in class_name.lower() for exc in excluded_classes)
                
                # Check title exclusions (for terminals showing script names)
                is_excluded = is_excluded or any(exc.lower() in title.lower() for exc in excluded_titles)
                
                if not is_excluded:
                    rect = win32gui.GetWindowRect(hwnd)
                    width = rect[2] - rect[0]
                    height = rect[3] - rect[1]
                    # Window must have reasonable size (not minimized)
                    if width > 200 and height > 200:
                        candidates.append({
                            'hwnd': hwnd,
                            'title': title,
                            'class': class_name,
                            'rect': rect,
                            'size': (width, height)
                        })
        return True  # Continue enumeration to find all candidates
    
    win32gui.EnumWindows(enum_callback, None)
    
    # Debug: show all candidates
    if candidates:
        print_debug(f"Found {len(candidates)} candidate window(s):")
        for i, c in enumerate(candidates):
            print_debug(f"  {i+1}. '{c['title']}' (class: {c['class']}, size: {c['size'][0]}x{c['size'][1]})")
    
    # Look for the actual FMRTE application
    # It should have a title like "FMRTE 26.0.6-build22" (with version number)
    for c in candidates:
        # FMRTE app window title starts with "FMRTE" and has version info
        if c['title'].startswith('FMRTE') and ('build' in c['title'].lower() or '26' in c['title']):
            hwnd = c['hwnd']
            left, top, right, bottom = c['rect']
            print_status(f"Found FMRTE window: '{c['title']}'")
            print_status(f"Window class: {c['class']}")
            print_status(f"Window rect: ({left}, {top}) to ({right}, {bottom})")
            print_status(f"Window size: {right-left}x{bottom-top}")
            return (hwnd, left, top, right, bottom)
    
    # Fallback: if no specific match, take first candidate
    if candidates:
        c = candidates[0]
        hwnd = c['hwnd']
        left, top, right, bottom = c['rect']
        print_status(f"Using first candidate: '{c['title']}'")
        print_status(f"Window class: {c['class']}")
        print_status(f"Window rect: ({left}, {top}) to ({right}, {bottom})")
        print_status(f"Window size: {right-left}x{bottom-top}")
        return (hwnd, left, top, right, bottom)
    
    return None


def activate_fmrte_window(hwnd):
    """
    Bring FMRTE window to foreground and focus it.
    
    Args:
        hwnd: Window handle
        
    Returns:
        bool: True if successful
    """
    try:
        # Restore if minimized
        if win32gui.IsIconic(hwnd):
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            time.sleep(ACTION_DELAY)
        
        # Bring to foreground
        try:
            win32gui.SetForegroundWindow(hwnd)
        except:
            pass
        
        time.sleep(ACTION_DELAY)
        return True
    except Exception as e:
        print_error(f"Failed to activate window: {e}")
        return False


def get_tab_coordinates(window_rect, tab_name):
    """
    Get screen coordinates for clicking on a specific tab.
    
    Based on calibration for FMRTE 26.0.6:
    - Tab Y position: 205 pixels from window top
    - Tab X positions from window left:
      - "Brixham": 189px
      - "Brixham U21": 284px
      - "Brixham U18": 385px
    
    Args:
        window_rect: (hwnd, left, top, right, bottom)
        tab_name: Name of the tab
        
    Returns:
        tuple: (x, y) screen coordinates
    """
    hwnd, left, top, right, bottom = window_rect
    
    # Y position: tabs are 205 pixels from window top
    y = top + 205
    
    # X positions for each tab (from window left edge)
    tab_x_offsets = {
        "Brixham": 189,
        "Brixham U21": 284,
        "Brixham U18": 385,
    }
    
    x_offset = tab_x_offsets.get(tab_name, 189)
    x = left + x_offset
    
    return (x, y)


def get_player_grid_coordinates(window_rect):
    """
    Get screen coordinates for clicking on the player grid (Name column, first data row).
    
    Based on calibration for FMRTE 26.0.6:
    - First player row X: 335 pixels from window left (in Name column)
    - First player row Y: 270 pixels from window top
    
    Args:
        window_rect: (hwnd, left, top, right, bottom)
        
    Returns:
        tuple: (x, y) screen coordinates
    """
    hwnd, left, top, right, bottom = window_rect
    
    # First player name cell position
    x = left + 335
    y = top + 270
    
    return (x, y)


def click_tab(window_rect, tab_name):
    """Click on a specific tab in FMRTE."""
    print_status(f"Switching to '{tab_name}' tab...")
    
    try:
        x, y = get_tab_coordinates(window_rect, tab_name)
        print_debug(f"Clicking tab at ({x}, {y})")
        
        pyautogui.click(x, y)
        time.sleep(ACTION_DELAY)
        
        print_status(f"Clicked '{tab_name}' tab")
        return True
        
    except Exception as e:
        print_error(f"Failed to click tab '{tab_name}': {e}")
        return False


def refresh_data():
    """Send Ctrl+R to refresh data in FMRTE."""
    print_status("Refreshing data (Ctrl+R)...")
    
    try:
        pyautogui.hotkey('ctrl', 'r')
        time.sleep(ACTION_DELAY * 3)  # Give time for refresh
        return True
    except Exception as e:
        print_error(f"Failed to refresh: {e}")
        return False


def copy_squad_data(window_rect, tab_name):
    """
    Copy data from the current squad tab in FMRTE.
    
    Strategy:
    1. Click on a player name in the grid to focus the data area
    2. Press Ctrl+A to select all players
    3. Press Ctrl+C to copy to clipboard
    """
    print_status(f"Copying data from '{tab_name}'...")
    
    # Clear clipboard first
    try:
        pyperclip.copy("")
    except:
        pass
    
    # Get player grid coordinates
    grid_x, grid_y = get_player_grid_coordinates(window_rect)
    
    # Click on the player grid area to focus it
    print_debug(f"Clicking player grid at ({grid_x}, {grid_y})")
    pyautogui.click(grid_x, grid_y)
    time.sleep(ACTION_DELAY)
    
    # Select all (Ctrl+A)
    print_status("Selecting all data (Ctrl+A)...")
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(ACTION_DELAY)
    
    # Copy (Ctrl+C)
    print_status("Copying to clipboard (Ctrl+C)...")
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(COPY_DELAY)  # Give clipboard time to populate
    
    # Get clipboard content
    clipboard_data = pyperclip.paste()
    
    if not clipboard_data or len(clipboard_data.strip()) == 0:
        # Try clicking lower in the grid
        print_status("First attempt failed, trying lower position...")
        
        grid_y_lower = grid_y + 30
        print_debug(f"Clicking at ({grid_x}, {grid_y_lower})")
        pyautogui.click(grid_x, grid_y_lower)
        time.sleep(ACTION_DELAY)
        
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(ACTION_DELAY)
        
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(COPY_DELAY)
        
        clipboard_data = pyperclip.paste()
        
        if not clipboard_data or len(clipboard_data.strip()) == 0:
            raise Exception("Clipboard is empty after copy operation - click may have missed the player grid")
    
    # Count rows for feedback
    row_count = len(clipboard_data.strip().split('\n'))
    print_status(f"Successfully copied {row_count} rows from '{tab_name}'")
    
    return clipboard_data


def parse_tsv_data(tsv_string):
    """Parse tab-separated values string into a list of rows, stripping quotes."""
    rows = []
    for line in tsv_string.strip().split('\n'):
        if line.strip():
            cells = line.split('\t')
            # Strip surrounding quotes from each cell
            cleaned_cells = []
            for cell in cells:
                cell = cell.strip()
                # Remove surrounding double quotes
                if cell.startswith('"') and cell.endswith('"'):
                    cell = cell[1:-1]
                # Handle escaped quotes inside ("" -> ")
                cell = cell.replace('""', '"')
                cleaned_cells.append(cell)
            rows.append(cleaned_cells)
    return rows


def write_to_excel(squad_data_dict):
    """Write squad data to Excel file."""
    print_divider()
    print_status(f"Opening Excel file: {EXCEL_FILE_PATH}")
    
    try:
        if not EXCEL_FILE_PATH.exists():
            raise Exception(f"Excel file not found: {EXCEL_FILE_PATH}")
        
        wb = load_workbook(EXCEL_FILE_PATH)
        
        if TARGET_SHEET not in wb.sheetnames:
            raise Exception(f"Sheet '{TARGET_SHEET}' not found in workbook")
        
        ws = wb[TARGET_SHEET]
        print_status(f"Accessing sheet: '{TARGET_SHEET}'")
        
        # Clear existing data (preserve row 1 headers)
        print_status("Clearing existing data (preserving headers in row 1)...")
        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
        
        # Start writing at row 2
        current_row = 2
        
        # Write data for each squad in order
        for squad_name in TABS:
            if squad_name not in squad_data_dict:
                print_error(f"No data found for '{squad_name}', skipping...")
                continue
            
            print_status(f"Writing '{squad_name}' data starting at row {current_row}...")
            
            rows = parse_tsv_data(squad_data_dict[squad_name])
            
            # FMRTE clipboard data has no header row - use all rows
            data_rows = rows
            
            if not data_rows:
                print_error(f"No data rows found for '{squad_name}'")
                continue
            
            for row_data in data_rows:
                for col_idx, cell_value in enumerate(row_data, start=1):
                    try:
                        if '.' in cell_value:
                            cell_value = float(cell_value)
                        elif cell_value.isdigit() or (cell_value.startswith('-') and cell_value[1:].isdigit()):
                            cell_value = int(cell_value)
                    except (ValueError, AttributeError):
                        pass
                    
                    ws.cell(row=current_row, column=col_idx, value=cell_value)
                
                current_row += 1
            
            print_status(f"Wrote {len(data_rows)} rows for '{squad_name}'")
        
        print_status("Saving Excel file...")
        wb.save(EXCEL_FILE_PATH)
        wb.close()
        
        print_status(f"Successfully saved {current_row - 2} total rows to Excel")
        
    except Exception as e:
        print_error(f"Failed to write to Excel: {e}")
        raise


def run_update_player_data():
    """Run the update_player_data.py script to generate CSV files."""
    print_divider()
    print_status("Running update_player_data.py to generate CSV files...")
    
    try:
        script_path = Path(__file__).parent / "update_player_data.py"
        
        if not script_path.exists():
            print_error(f"Script not found: {script_path}")
            return False
        
        result = subprocess.run(
            [sys.executable, str(script_path)],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0:
            print_status("Successfully generated CSV files")
            if result.stdout:
                print(result.stdout)
            return True
        else:
            print_error(f"update_player_data.py failed with exit code {result.returncode}")
            if result.stderr:
                print(result.stderr)
            return False
    
    except subprocess.TimeoutExpired:
        print_error("update_player_data.py timed out after 30 seconds")
        return False
    except Exception as e:
        print_error(f"Failed to run update_player_data.py: {e}")
        return False


def calibrate_positions(window_rect):
    """Interactive calibration mode."""
    hwnd, left, top, right, bottom = window_rect
    window_width = right - left
    window_height = bottom - top
    
    print_divider()
    print("CALIBRATION MODE")
    print_divider()
    print()
    print(f"FMRTE Window: ({left}, {top}) to ({right}, {bottom})")
    print(f"Window size: {window_width}x{window_height}")
    print()
    print("Move your mouse over the FMRTE window and watch the coordinates.")
    print("The 'Window offset' values are what you need for the config.")
    print()
    print("Press Ctrl+C to exit.")
    print()
    
    try:
        while True:
            x, y = pyautogui.position()
            offset_x = x - left
            offset_y = y - top
            
            in_window = (0 <= offset_x <= window_width and 0 <= offset_y <= window_height)
            status = "IN FMRTE" if in_window else "OUTSIDE"
            
            print(f"Screen: ({x:4d}, {y:4d}) | Window offset: ({offset_x:4d}, {offset_y:4d}) [{status}]    ", end='\r')
            time.sleep(0.05)
    except KeyboardInterrupt:
        print()
        print()
        print_divider()
        print("Calibration ended.")


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(description="FMRTE to Excel Automation")
    parser.add_argument("--calibrate", action="store_true", help="Calibration mode")
    parser.add_argument("--skip-refresh", action="store_true", help="Skip Ctrl+R refresh")
    parser.add_argument("--skip-update", action="store_true", help="Skip update_player_data.py")
    parser.add_argument("--test-tab", choices=TABS, help="Test click on a specific tab")
    parser.add_argument("--test-grid", action="store_true", help="Test click on player grid")
    args = parser.parse_args()
    
    print_divider()
    print("FMRTE to Excel Automation")
    print("FM26 Lineup Optimizer - Data Import Tool")
    print_divider()
    print()
    print("SAFETY: Move mouse to any screen corner to abort.")
    print()
    
    try:
        # Find FMRTE window
        window_rect = find_fmrte_window()
        
        if window_rect is None:
            print_error("Could not find FMRTE window!")
            print_error("Make sure FMRTE is running and not minimized.")
            return 1
        
        hwnd = window_rect[0]
        
        # Activate window
        if not activate_fmrte_window(hwnd):
            print_error("Could not activate FMRTE window")
            return 1
        
        time.sleep(ACTION_DELAY)
        
        # Calibration mode
        if args.calibrate:
            calibrate_positions(window_rect)
            return 0
        
        # Test tab click
        if args.test_tab:
            print_status(f"Test clicking on '{args.test_tab}' tab...")
            x, y = get_tab_coordinates(window_rect, args.test_tab)
            print_status(f"Target coordinates: ({x}, {y})")
            
            # Show current mouse position
            point = ctypes.wintypes.POINT()
            ctypes.windll.user32.GetCursorPos(ctypes.byref(point))
            print_status(f"Current mouse position: ({point.x}, {point.y})")
            
            # Show virtual screen info
            vscreen_x = ctypes.windll.user32.GetSystemMetrics(76)
            vscreen_y = ctypes.windll.user32.GetSystemMetrics(77)
            vscreen_w = ctypes.windll.user32.GetSystemMetrics(78)
            vscreen_h = ctypes.windll.user32.GetSystemMetrics(79)
            print_status(f"Virtual screen: origin ({vscreen_x}, {vscreen_y}), size {vscreen_w}x{vscreen_h}")
            
            print_status("Moving mouse to target in 3 seconds...")
            for i in range(3, 0, -1):
                print(f"  {i}...")
                time.sleep(1)
            
            print_status(f"Moving mouse using raw API...")
            new_x, new_y = move_mouse_raw(x, y)
            print_status(f"Mouse now at: ({new_x}, {new_y})")
            
            if abs(new_x - x) > 20 or abs(new_y - y) > 20:
                print_error(f"Mouse did not move to target! Expected ({x}, {y}), got ({new_x}, {new_y})")
            else:
                print_status("Mouse moved successfully!")
            
            time.sleep(1)
            print_status("Clicking...")
            click_raw()
            print_status("Clicked! Check if the correct tab was selected.")
            return 0
        
        # Test grid click
        if args.test_grid:
            print_status("Test clicking on player grid...")
            x, y = get_player_grid_coordinates(window_rect)
            print_status(f"Target coordinates: ({x}, {y})")
            
            # Show current mouse position
            current_x, current_y = pyautogui.position()
            print_status(f"Current mouse position: ({current_x}, {current_y})")
            
            print_status("Moving mouse to target in 3 seconds...")
            for i in range(3, 0, -1):
                print(f"  {i}...")
                time.sleep(1)
            
            print_status(f"Moving mouse from ({current_x}, {current_y}) to ({x}, {y})...")
            
            # Try pyautogui first
            pyautogui.moveTo(x, y, duration=0.5)
            
            new_x, new_y = pyautogui.position()
            print_status(f"Mouse now at: ({new_x}, {new_y})")
            
            # If pyautogui didn't work, try win32api directly
            if abs(new_x - x) > 10 or abs(new_y - y) > 10:
                print_status("pyautogui moveTo failed, trying win32api...")
                win32api.SetCursorPos((x, y))
                new_x, new_y = win32api.GetCursorPos()
                print_status(f"Mouse now at (via win32api): ({new_x}, {new_y})")
            
            time.sleep(1)
            print_status("Clicking...")
            pyautogui.click()
            print_status("Clicked! Check if a player was selected.")
            return 0
        
        # Copy data from each squad tab
        squad_data = {}
        
        for tab_name in TABS:
            print_divider()
            
            if not click_tab(window_rect, tab_name):
                print_error(f"Failed to click '{tab_name}' tab")
                user_input = input(f"Press Enter after manually switching to '{tab_name}' tab (or 'q' to quit): ")
                if user_input.lower() == 'q':
                    return 1
            
            time.sleep(ACTION_DELAY * 2)
            
            if not args.skip_refresh:
                refresh_data()
            
            try:
                data = copy_squad_data(window_rect, tab_name)
                squad_data[tab_name] = data
            except Exception as e:
                print_error(f"Failed to copy data from '{tab_name}': {e}")
                user_input = input("Press Enter to continue (or 'q' to quit): ")
                if user_input.lower() == 'q':
                    return 1
        
        if not squad_data:
            print_error("No squad data was copied.")
            return 1
        
        print_divider()
        print_status(f"Successfully copied data from {len(squad_data)} squad(s)")
        
        # Write to Excel
        write_to_excel(squad_data)
        
        # Run update script
        if not args.skip_update:
            run_update_player_data()
        
        print_divider()
        print("SUCCESS! FMRTE data has been imported.")
        print_divider()
        
        return 0
        
    except pyautogui.FailSafeException:
        print()
        print_error("ABORTED! Mouse moved to screen corner.")
        return 1
    except KeyboardInterrupt:
        print()
        print_status("Operation cancelled by user")
        return 1
    except Exception as e:
        print_error(f"Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
