"""
FMRTE to Excel Automation Script
---------------------------------
Automates the process of copying player data from Football Manager Real Time Editor (FMRTE)
into an Excel spreadsheet for the FM26 Lineup Optimizer.

Author: Doug Mason (2025)
License: MIT
"""

import time
import subprocess
import sys
import argparse
from pathlib import Path

try:
    from pywinauto import Application, Desktop
    from pywinauto.keyboard import send_keys
    from pywinauto.timings import wait_until
    import pywinauto
except ImportError:
    print("ERROR: pywinauto not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    import pyperclip
except ImportError:
    print("ERROR: pyperclip not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install -r requirements.txt")
    sys.exit(1)


# Configuration
EXCEL_FILE_PATH = Path(r"C:\GitHub\fm26-lineup-optimizer\FM26 Players.xlsx")
TARGET_SHEET = "Paste Full"
FMRTE_WINDOW_TITLE = "FMRTE"
TABS = ["Brixham", "Brixham U21", "Brixham U18"]
ACTION_DELAY = 0.8  # Seconds to wait between actions


def print_divider():
    """Print a visual divider for terminal output."""
    print("=" * 70)


def print_status(message):
    """Print a status message."""
    print(f"[STATUS] {message}")


def print_error(message):
    """Print an error message."""
    print(f"[ERROR] {message}")


def print_debug(message):
    """Print a debug message."""
    print(f"[DEBUG] {message}")


def list_all_windows():
    """
    List all detectable windows for debugging.
    Tries both UIA and Win32 backends.
    """
    print_divider()
    print("DEBUG MODE: Listing all detectable windows")
    print_divider()

    for backend in ["uia", "win32"]:
        print(f"\n{backend.upper()} Backend:")
        print("-" * 70)

        try:
            desktop = Desktop(backend=backend)
            windows = desktop.windows()

            if not windows:
                print(f"  No windows found with {backend} backend")
                continue

            for idx, window in enumerate(windows, 1):
                try:
                    title = window.window_text()
                    class_name = window.class_name()
                    is_visible = window.is_visible()
                    is_enabled = window.is_enabled()

                    if title or "FMRTE" in class_name.upper():
                        print(f"  [{idx}] Title: '{title}'")
                        print(f"      Class: {class_name}")
                        print(f"      Visible: {is_visible}, Enabled: {is_enabled}")
                        print()
                except Exception as e:
                    continue

        except Exception as e:
            print(f"  ERROR with {backend} backend: {e}")
            print()

    print_divider()


def get_foreground_window_after_delay(delay=3):
    """
    Wait for user to click a window, then return that window.

    Args:
        delay: Seconds to wait before capturing foreground window

    Returns:
        Window object or None
    """
    print_status(f"Please click on the FMRTE window in the next {delay} seconds...")

    for i in range(delay, 0, -1):
        print(f"  {i}...", end="\r")
        time.sleep(1)

    print("  Capturing window...    ")

    try:
        # Try both backends
        for backend in ["win32", "uia"]:
            try:
                app = Application(backend=backend).connect(active_only=True)
                window = app.top_window()
                title = window.window_text()
                print_status(f"Captured window: '{title}' (backend: {backend})")
                return window
            except Exception:
                continue

        print_error("Could not capture foreground window")
        return None

    except Exception as e:
        print_error(f"Failed to capture foreground window: {e}")
        return None


def is_likely_fmrte_window(window):
    """
    Check if a window is likely to be FMRTE (not a terminal/IDE/browser).

    Args:
        window: pywinauto window object

    Returns:
        bool: True if likely FMRTE, False otherwise
    """
    try:
        title = window.window_text()
        class_name = window.class_name()

        # Exclude terminal/IDE/browser windows by class name
        excluded_classes = [
            "WindowsTerminal",  # Windows Terminal
            "ConsoleWindowClass",  # Old cmd windows
            "mintty",  # Git Bash, Cygwin
            "Chrome_WidgetWin",  # Chrome/Edge
            "MozillaWindowClass",  # Firefox
            "ApplicationFrameWindow",  # Modern UWP apps
            "Notepad",  # Notepad
            "PX_WINDOW_CLASS",  # Visual Studio Code
            "SunAwtFrame",  # Java apps
            "Qt5",  # Qt-based apps (many IDEs)
        ]

        # Check if class name contains any excluded patterns
        for excluded in excluded_classes:
            if excluded.lower() in class_name.lower():
                return False

        # Prioritize windows that start with "FMRTE" (exact match at beginning)
        if title and title.upper().startswith("FMRTE"):
            return True

        # If title contains "FMRTE" but not at start, be more cautious
        # Check that it's not a terminal tab name
        if title and "FMRTE" in title.upper():
            # Exclude if it looks like a tab/document name pattern
            # E.g., "FMRTE Window Detection Issue" in a terminal
            # FMRTE real windows are usually just "FMRTE version-build"
            # Count spaces - real FMRTE has 1-2 spaces max
            space_count = title.count(' ')
            if space_count > 2:
                print_debug(f"Excluding '{title}' - too many spaces (likely terminal tab)")
                return False

            # Check for common terminal/IDE indicators in title
            terminal_indicators = [
                "issue", "detection", "window", "tab", "file", "editor",
                "bash", "powershell", "terminal", "wsl", "command"
            ]
            title_lower = title.lower()
            for indicator in terminal_indicators:
                if indicator in title_lower:
                    print_debug(f"Excluding '{title}' - contains terminal indicator '{indicator}'")
                    return False

            return True

        return False

    except Exception:
        return False


def find_fmrte_window():
    """
    Find and connect to the FMRTE window.
    Tries multiple detection methods and backends.

    Returns:
        Window object for FMRTE

    Raises:
        Exception: If FMRTE window cannot be found
    """
    print_status("Searching for FMRTE window...")

    fmrte_window = None

    # Method 1: Try both backends with partial title matching + filtering
    for backend in ["win32", "uia"]:
        print_status(f"Trying {backend.upper()} backend...")

        try:
            desktop = Desktop(backend=backend)
            windows = desktop.windows()

            for window in windows:
                try:
                    title = window.window_text()
                    # Case-insensitive partial match
                    if title and FMRTE_WINDOW_TITLE.upper() in title.upper():
                        # Check if window is visible and enabled
                        if window.is_visible() and window.is_enabled():
                            # Apply filtering to exclude terminals/IDEs
                            if is_likely_fmrte_window(window):
                                fmrte_window = window
                                print_status(f"Found FMRTE window: '{title}' (backend: {backend})")
                                break
                            else:
                                print_debug(f"Skipped '{title}' - filtered out")
                except Exception:
                    continue

            if fmrte_window:
                break

        except Exception as e:
            print_debug(f"{backend} backend failed: {e}")
            continue

    # Method 2: Try connecting by process name
    if fmrte_window is None:
        print_status("Trying to connect by process name 'FMRTE.exe'...")

        for backend in ["win32", "uia"]:
            try:
                app = Application(backend=backend).connect(path="FMRTE.exe", timeout=2)
                fmrte_window = app.top_window()
                title = fmrte_window.window_text()
                print_status(f"Connected to FMRTE via process: '{title}' (backend: {backend})")
                break
            except Exception:
                continue

    # Method 3: Try exact title matches for known FMRTE versions
    if fmrte_window is None:
        print_status("Trying exact title matches for common FMRTE versions...")

        known_titles = [
            "FMRTE 26.0.5-build21",
            "FMRTE 26",
            "FMRTE",
        ]

        for backend in ["win32", "uia"]:
            for known_title in known_titles:
                try:
                    desktop = Desktop(backend=backend)
                    fmrte_window = desktop.window(title=known_title)
                    if fmrte_window.exists():
                        print_status(f"Found FMRTE with exact title: '{known_title}' (backend: {backend})")
                        break
                except Exception:
                    continue

            if fmrte_window:
                break

    # Method 4: Manual window selection fallback
    if fmrte_window is None:
        print_error("Could not automatically detect FMRTE window")
        print_status("Falling back to manual window selection...")
        print()

        response = input("Do you want to manually select the FMRTE window? (y/n): ")

        if response.lower() == 'y':
            fmrte_window = get_foreground_window_after_delay(delay=3)

    # Final check
    if fmrte_window is None:
        raise Exception(
            "Could not find FMRTE window using any detection method.\n"
            "       Troubleshooting:\n"
            "         1. Make sure FMRTE is running and visible\n"
            "         2. Try running: python fmrte_to_excel.py --debug\n"
            "         3. Check if FMRTE window title contains 'FMRTE'"
        )

    # Bring window to foreground
    try:
        if not fmrte_window.is_active():
            fmrte_window.set_focus()
            time.sleep(ACTION_DELAY)
    except Exception as e:
        print_debug(f"Could not focus window: {e}")
        # Continue anyway, window might still be usable

    return fmrte_window


def debug_window_controls(fmrte_window):
    """
    Print all child controls of the FMRTE window for debugging.

    Args:
        fmrte_window: The FMRTE window object
    """
    print_divider()
    print("DEBUG: FMRTE Window Control Structure")
    print_divider()

    try:
        print(f"Main Window: '{fmrte_window.window_text()}'")
        print(f"Class Name: {fmrte_window.class_name()}")
        print()

        # Get all descendants
        print("Searching for tab-related controls...")
        print("-" * 70)

        # Look for controls that might be tabs
        tab_keywords = ["brixham", "u21", "u18", "tab"]

        descendants = fmrte_window.descendants()
        found_count = 0

        for control in descendants:
            try:
                title = control.window_text()
                control_type = control.element_info.control_type if hasattr(control, 'element_info') else "Unknown"
                class_name = control.class_name()

                # Filter to show only potentially relevant controls
                title_lower = title.lower() if title else ""
                is_relevant = any(keyword in title_lower for keyword in tab_keywords)

                if is_relevant or not title:
                    # Always show controls that match our keywords, or have no title (might be containers)
                    continue

                if is_relevant:
                    print(f"  [{found_count + 1}] Title: '{title}'")
                    print(f"      Type: {control_type}")
                    print(f"      Class: {class_name}")
                    try:
                        rect = control.rectangle()
                        print(f"      Position: ({rect.left}, {rect.top}) Size: ({rect.width()}x{rect.height()})")
                    except:
                        pass
                    print()
                    found_count += 1

            except Exception:
                continue

        if found_count == 0:
            print("  No tab-related controls found by keyword search.")
            print("  Showing ALL controls with text...")
            print()

            for idx, control in enumerate(descendants[:50], 1):  # Limit to first 50
                try:
                    title = control.window_text()
                    if title and len(title) > 0:
                        control_type = control.element_info.control_type if hasattr(control, 'element_info') else "Unknown"
                        class_name = control.class_name()
                        print(f"  [{idx}] Title: '{title}'")
                        print(f"      Type: {control_type}, Class: {class_name}")
                        print()
                except Exception:
                    continue

        print_divider()

    except Exception as e:
        print_error(f"Failed to list window controls: {e}")
        print_divider()


def click_tab(fmrte_window, tab_name):
    """
    Click on a specific tab in FMRTE.
    Uses .click() instead of .click_input() to avoid cursor positioning permission issues.

    Args:
        fmrte_window: The FMRTE window object
        tab_name: Name of the tab to click

    Returns:
        bool: True if successful, False otherwise
    """
    print_status(f"Switching to '{tab_name}' tab...")

    try:
        # Try to find and click the tab control
        # FMRTE may use different control types, so we try multiple approaches
        # Using .click() instead of .click_input() to avoid admin requirements

        # Approach 1: Try finding tab by name directly (TabItem)
        try:
            tab = fmrte_window.child_window(title=tab_name, control_type="TabItem")
            tab.click()
            time.sleep(ACTION_DELAY)
            print_status(f"Successfully clicked '{tab_name}' tab (TabItem)")
            return True
        except Exception as e:
            print_debug(f"TabItem approach failed: {e}")

        # Approach 2: Try finding as a button
        try:
            tab = fmrte_window.child_window(title=tab_name, control_type="Button")
            tab.click()
            time.sleep(ACTION_DELAY)
            print_status(f"Successfully clicked '{tab_name}' button")
            return True
        except Exception as e:
            print_debug(f"Button approach failed: {e}")

        # Approach 3: Try finding as Text control (some tabs are text labels)
        try:
            tab = fmrte_window.child_window(title=tab_name, control_type="Text")
            tab.click()
            time.sleep(ACTION_DELAY)
            print_status(f"Successfully clicked '{tab_name}' text")
            return True
        except Exception as e:
            print_debug(f"Text control approach failed: {e}")

        # Approach 4: Try finding with best_match (any control type)
        try:
            tab = fmrte_window.child_window(title=tab_name, found_index=0)
            tab.click()
            time.sleep(ACTION_DELAY)
            print_status(f"Successfully clicked '{tab_name}' (best match)")
            return True
        except Exception as e:
            print_debug(f"Best match approach failed: {e}")

        # Approach 5: Try with partial matching
        try:
            tab = fmrte_window.child_window(title_re=f".*{tab_name}.*", found_index=0)
            tab.click()
            time.sleep(ACTION_DELAY)
            print_status(f"Successfully clicked '{tab_name}' (partial match)")
            return True
        except Exception as e:
            print_debug(f"Partial match approach failed: {e}")

        print_error(f"Could not find '{tab_name}' tab using any automated method")
        print_status("Try running: python fmrte_to_excel.py --debug-tabs")
        return False

    except Exception as e:
        print_error(f"Failed to click tab '{tab_name}': {e}")
        return False


def copy_squad_data(fmrte_window, tab_name):
    """
    Copy data from a squad tab in FMRTE.
    Uses keyboard-only approach to avoid cursor positioning permission issues.

    Args:
        fmrte_window: The FMRTE window object
        tab_name: Name of the tab to copy from

    Returns:
        str: Clipboard content (tab-separated values)

    Raises:
        Exception: If data cannot be copied
    """
    print_status(f"Copying data from '{tab_name}'...")

    try:
        # Clear clipboard before copying
        pyperclip.copy("")

        # Ensure FMRTE window has focus (don't click specific controls)
        try:
            if not fmrte_window.is_active():
                fmrte_window.set_focus()
                time.sleep(ACTION_DELAY)
        except Exception as e:
            print_debug(f"Could not set focus: {e}")

        # Keyboard-only approach: Tab to navigate to grid, then select all
        # First, try clicking the grid without cursor positioning
        try:
            grid = fmrte_window.child_window(control_type="DataGrid", found_index=0)
            # Use .click() instead of .click_input() - doesn't require cursor positioning
            grid.click()
            time.sleep(ACTION_DELAY / 2)
            print_status("Found and focused on DataGrid")
        except Exception as e:
            print_debug(f"Could not click DataGrid: {e}")
            # Try focusing on the window and using Tab key to navigate to grid
            print_status("Using Tab key to navigate to data grid...")
            send_keys("{TAB}")
            time.sleep(ACTION_DELAY / 2)

        # Select all (Ctrl+A)
        print_status("Selecting all data (Ctrl+A)...")
        send_keys("^a")
        time.sleep(ACTION_DELAY)

        # Copy (Ctrl+C)
        print_status("Copying to clipboard (Ctrl+C)...")
        send_keys("^c")
        time.sleep(ACTION_DELAY * 1.5)  # Give more time for large datasets

        # Get clipboard content
        clipboard_data = pyperclip.paste()

        if not clipboard_data or len(clipboard_data.strip()) == 0:
            raise Exception("Clipboard is empty after copy operation")

        # Count rows for feedback
        row_count = len(clipboard_data.strip().split('\n'))
        print_status(f"Successfully copied {row_count} rows from '{tab_name}'")

        return clipboard_data

    except Exception as e:
        print_error(f"Failed to copy data from '{tab_name}': {e}")
        # Check if it's a permission error
        if "rights" in str(e).lower() or "setcursorpos" in str(e).lower():
            print_error("PERMISSION ERROR: This script needs administrator rights")
            print_error("Please RIGHT-CLICK fmrte_update.bat and select 'Run as administrator'")
        raise


def parse_tsv_data(tsv_string):
    """
    Parse tab-separated values string into a list of rows.

    Args:
        tsv_string: Tab-separated values as string

    Returns:
        list: List of lists representing rows and cells
    """
    rows = []
    for line in tsv_string.strip().split('\n'):
        if line.strip():  # Skip empty lines
            cells = line.split('\t')
            rows.append(cells)
    return rows


def write_to_excel(squad_data_dict):
    """
    Write squad data to Excel file.

    Args:
        squad_data_dict: Dictionary with squad names as keys and TSV data as values

    Raises:
        Exception: If Excel file cannot be written
    """
    print_divider()
    print_status(f"Opening Excel file: {EXCEL_FILE_PATH}")

    try:
        if not EXCEL_FILE_PATH.exists():
            raise Exception(f"Excel file not found: {EXCEL_FILE_PATH}")

        # Load workbook
        wb = load_workbook(EXCEL_FILE_PATH)

        if TARGET_SHEET not in wb.sheetnames:
            raise Exception(f"Sheet '{TARGET_SHEET}' not found in workbook")

        ws = wb[TARGET_SHEET]
        print_status(f"Accessing sheet: '{TARGET_SHEET}'")

        # Clear existing data (preserve row 1 headers)
        print_status("Clearing existing data (preserving headers in row 1)...")
        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)

        # Start writing at row 2 (row 1 is headers)
        current_row = 2

        # Write data for each squad in order
        for squad_name in TABS:
            if squad_name not in squad_data_dict:
                print_error(f"No data found for '{squad_name}', skipping...")
                continue

            print_status(f"Writing '{squad_name}' data starting at row {current_row}...")

            # Parse the TSV data
            rows = parse_tsv_data(squad_data_dict[squad_name])

            # Skip header row from FMRTE data (we preserve Excel headers in row 1)
            data_rows = rows[1:] if len(rows) > 1 else []

            if not data_rows:
                print_error(f"No data rows found for '{squad_name}'")
                continue

            # Write each row
            for row_data in data_rows:
                for col_idx, cell_value in enumerate(row_data, start=1):
                    # Try to convert to number if possible
                    try:
                        if '.' in cell_value:
                            cell_value = float(cell_value)
                        elif cell_value.isdigit():
                            cell_value = int(cell_value)
                    except (ValueError, AttributeError):
                        pass  # Keep as string

                    ws.cell(row=current_row, column=col_idx, value=cell_value)

                current_row += 1

            print_status(f"Wrote {len(data_rows)} rows for '{squad_name}'")

        # Save workbook
        print_status("Saving Excel file...")
        wb.save(EXCEL_FILE_PATH)
        wb.close()

        print_status(f"Successfully saved {current_row - 2} total rows to Excel")

    except Exception as e:
        print_error(f"Failed to write to Excel: {e}")
        raise


def run_update_player_data():
    """
    Run the update_player_data.py script to generate CSV files.

    Returns:
        bool: True if successful, False otherwise
    """
    print_divider()
    print_status("Running update_player_data.py to generate CSV files...")

    try:
        script_path = Path(__file__).parent / "update_player_data.py"

        if not script_path.exists():
            print_error(f"Script not found: {script_path}")
            print_error("Please run update_player_data.py manually to generate CSV files")
            return False

        result = subprocess.run(
            [sys.executable, str(script_path)],
            capture_output=True,
            text=True,
            timeout=30
        )

        if result.returncode == 0:
            print_status("Successfully generated CSV files")
            if result.stdout:
                print(result.stdout)
            return True
        else:
            print_error(f"update_player_data.py failed with exit code {result.returncode}")
            if result.stderr:
                print(result.stderr)
            return False

    except subprocess.TimeoutExpired:
        print_error("update_player_data.py timed out after 30 seconds")
        return False
    except Exception as e:
        print_error(f"Failed to run update_player_data.py: {e}")
        return False


def main():
    """Main execution function."""
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description="FMRTE to Excel Automation - Automates copying player data from FMRTE to Excel"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug mode to list all detectable windows"
    )
    parser.add_argument(
        "--debug-tabs",
        action="store_true",
        help="Debug tab controls - shows all child controls of FMRTE window"
    )
    args = parser.parse_args()

    print_divider()
    print("FMRTE to Excel Automation")
    print("FM26 Lineup Optimizer - Data Import Tool")
    print_divider()

    # Debug mode - list all windows and exit
    if args.debug:
        list_all_windows()
        print("\nDebug information displayed above.")
        print("Please check if FMRTE window is listed with its exact title.")
        print("\nIf you see FMRTE in the list, the script should work normally.")
        print("If not, try running FMRTE as administrator or check if it's minimized.")
        return 0

    # Debug tabs mode - show window controls and exit
    if args.debug_tabs:
        try:
            fmrte_window = find_fmrte_window()
            debug_window_controls(fmrte_window)
            print("\nDebug information displayed above.")
            print("Look for controls with 'Brixham', 'U21', or 'U18' in the title.")
            print("Note the control Type and Class to help fix tab detection.")
            return 0
        except Exception as e:
            print_error(f"Failed to debug tabs: {e}")
            return 1

    try:
        # Step 1: Find FMRTE window
        fmrte_window = find_fmrte_window()
        time.sleep(ACTION_DELAY)

        # Step 2: Copy data from each squad tab
        squad_data = {}

        for tab_name in TABS:
            print_divider()

            # Click the tab
            if not click_tab(fmrte_window, tab_name):
                print_error(f"Failed to switch to '{tab_name}' tab")
                user_input = input(f"Press Enter after manually switching to '{tab_name}' tab (or 'q' to quit): ")
                if user_input.lower() == 'q':
                    print_status("User cancelled operation")
                    return 1

            # Copy the data
            try:
                data = copy_squad_data(fmrte_window, tab_name)
                squad_data[tab_name] = data
            except Exception as e:
                print_error(f"Failed to copy data from '{tab_name}': {e}")
                user_input = input("Press Enter to continue with next squad (or 'q' to quit): ")
                if user_input.lower() == 'q':
                    print_status("User cancelled operation")
                    return 1

        # Step 3: Verify we have data
        if not squad_data:
            print_error("No squad data was copied. Exiting.")
            return 1

        print_divider()
        print_status(f"Successfully copied data from {len(squad_data)} squad(s)")

        # Step 4: Write to Excel
        write_to_excel(squad_data)

        # Step 5: Run update_player_data.py
        update_success = run_update_player_data()

        # Final status
        print_divider()
        if update_success:
            print("SUCCESS! FMRTE data has been imported and CSV files generated.")
            print("You can now run the team selector scripts.")
        else:
            print("PARTIAL SUCCESS! FMRTE data imported to Excel.")
            print("Please run 'python update_player_data.py' manually to generate CSVs.")
        print_divider()

        return 0

    except KeyboardInterrupt:
        print()
        print_status("Operation cancelled by user")
        return 1
    except Exception as e:
        print_divider()
        print_error(f"Unexpected error: {e}")
        print_divider()
        return 1


if __name__ == "__main__":
    sys.exit(main())
